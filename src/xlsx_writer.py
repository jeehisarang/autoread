"""인식된 행동을 엑셀(.xlsx) 파일에 누적 저장한다.

한 건이 기록될 때마다 즉시 저장하여, 프로그램이 중간에 종료되어도
기록이 최대한 유실되지 않도록 한다.
"""
import os
import threading

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS = ["시간", "게임 번호", "스트리트", "포지션", "행동", "금액", "내 행동", "원문(OCR)"]
AMOUNT_COL = HEADERS.index("금액") + 1  # openpyxl은 1부터 시작
# task.md "내 행동 구분 표시" - 빨간색 + 굵게로 강조(기존엔 초록색이었음).
HERO_FONT = Font(color="C00000", bold=True)


class XlsxLogger:
    def __init__(self, path: str):
        self.path = path
        self._lock = threading.Lock()
        self._wb = None
        self._ws = None
        self.last_error: str | None = None
        self.last_row: int | None = None
        self._open_or_create()

    def _open_or_create(self) -> None:
        if os.path.exists(self.path):
            try:
                self._wb = openpyxl.load_workbook(self.path)
                self._ws = self._wb.active
                current = [c.value for c in self._ws[1]] if self._ws.max_row >= 1 else []
                if current != HEADERS:
                    # 예전 버전 형식의 파일이면, 기존 데이터는 그대로 아래에 남겨두고
                    # 맨 위에 새 헤더 행을 추가해서 새 형식으로 이어서 사용한다.
                    self._ws.insert_rows(1)
                    for col, header in enumerate(HEADERS, start=1):
                        self._ws.cell(row=1, column=col, value=header)
                    self._save()
                return
            except Exception:
                pass  # 손상되었거나 열 수 없으면 새로 만든다.

        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._ws.title = "홀덤 기록"
        for col, header in enumerate(HEADERS, start=1):
            self._ws.cell(row=1, column=col, value=header)
            self._ws.column_dimensions[get_column_letter(col)].width = 14
        self._ws.column_dimensions[get_column_letter(len(HEADERS))].width = 60
        self._save()

    def _save(self) -> bool:
        try:
            self._wb.save(self.path)
            self.last_error = None
            return True
        except PermissionError:
            self.last_error = "파일이 다른 프로그램(엑셀 등)에서 열려 있어 저장할 수 없습니다. 파일을 닫아주세요."
            return False
        except Exception as e:
            self.last_error = f"저장 중 오류: {e}"
            return False

    def append(self, row: list, highlight: bool = False) -> bool:
        with self._lock:
            self._ws.append(row)
            last_row = self._ws.max_row
            self.last_row = last_row  # task.md "베팅 금액 보완" - 나중에 금액을 채워 넣을 때 씀
            if highlight:
                for col in range(1, len(row) + 1):
                    self._ws.cell(row=last_row, column=col).font = HERO_FONT
            return self._save()

    def update_amount(self, row: int, amount: str) -> bool:
        """행동을 기록할 때는 금액을 못 읽었지만, 몇 프레임 뒤(칩 숫자가 화면에
        남아있는 동안) 다시 읽는 데 성공하면 그 행 "금액" 칸만 채워 넣는다."""
        with self._lock:
            self._ws.cell(row=row, column=AMOUNT_COL, value=amount)
            return self._save()

    def row_count(self) -> int:
        with self._lock:
            return max(0, self._ws.max_row - 1)
